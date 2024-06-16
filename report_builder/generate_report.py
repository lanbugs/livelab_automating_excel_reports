from pymongo import MongoClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import sys

# Settings
username = 'root'
password = 'example'
host = 'localhost'
port = 27117
connection_string = f'mongodb://{username}:{password}@{host}:{port}/'



def optimize_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # add title line
    title = f"MAC Export - Date: {datetime.now().strftime('%Y-%m-%d')}"

    ws.insert_rows(1)
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=15)

    # Color headline
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    # Set text bold of filter line
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # enable filterline
    ws.auto_filter.ref = f"A2:{ws.dimensions.split(':')[1]}"

    # determine correct column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # freeze title and filter line
    ws.freeze_panes = ws['A3']

    # save modifications
    wb.save(file_path)


def main(doc):
    # connect to mongo db
    client = MongoClient(connection_string)
    db = client['mac_records']
    collection = db['records']

    # get all documents from collection
    documents = list(collection.find())

    # build dataframe
    df = pd.DataFrame(documents)

    # remove _id column
    df = df.drop(columns=['_id'])

    # convert list to comma seperated string
    df["destination_port"] = df["destination_port"].apply(lambda x: ', '.join(map(str, x)))

    # generate excel file
    df.to_excel(doc, index=False)

    # optimize it
    optimize_excel(doc)


if __name__ == '__main__':
    main(sys.argv[1])